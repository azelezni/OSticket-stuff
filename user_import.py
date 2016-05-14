from openpyxl import load_workbook
from ldap3 import Connection
import json
import time

"""
    Import users from Active Directory into osticket database
    search is dependent upon division and campus
"""


def mobile_parse(phone):
    """
    Parse mobile number and return in a predefined format: 052-123-4567
    :param phone: phone number
    :return: number in format xxx-xxxxxx
    """
    chars = ['+', '/', ' ', '-']
    mobile = phone
    for c in chars:
        mobile = mobile.replace(c, '')
    if mobile[0:3] == '972':  # replace only first 3 letters (in case there is a 972 in the actual number)
        mobile = mobile[3:]
        if mobile[0] != '0':  # check if there is a 0 after 972 (ie +972-052-123-4567)
            mobile = '0' + mobile
    mobile = mobile[0:3] + '-' + mobile[3:6] + '-' + mobile[6:]
    return mobile

users = {}
user_info = {}
ad_info = {}

''' set AD base, filter, and search attributes '''
division = ''
campus = ''
ad_base = ''
ad_filter = ''
ad_attributes = []

''' connect to AD '''
ad_conn = Connection('', user='', password='', auto_bind=True)
ad_conn.search(ad_base, ad_filter, attributes=ad_attributes)
do = len(ad_conn.entries)

for x in range(0, do):  # store all results from Active Directory
    ad_info[x] = json.loads(ad_conn.entries[x].entry_to_json())

''' store user information '''
for x in range(0, do):

    info = ad_info[x]

    user_info['name'] = ''.join(info['attributes']['cn'])
    user_info['mail'] = ''.join(info['attributes']['mail'])
    user_info['id'] = ''.join(info['attributes']['employeeID'])
    user_info['job_description'] = ''.join(info['attributes']['jobDescription'])
    user_info['manager_id'] = ''.join(info['attributes']['mgrID'])
    try:
        user_info['mobile'] = mobile_parse(''.join(info['attributes']['mobile']))
    except KeyError:
        pass

    ''' change filter and attributes to grab user's manager information '''
    ad_attributes = ['mail', 'name']
    ad_filter = '(&(objectcategory=person)(objectclass=user)(employeeID=%s))' % user_info['manager_id']
    ad_conn.search(ad_base, ad_filter, attributes=ad_attributes)
    manager_info = json.loads(ad_conn.entries[0].entry_to_json())

    ''' store manager information '''
    try:
        user_info['manager_name'] = ''.join(manager_info['attributes']['name'])
        user_info['manager_mail'] = ''.join(manager_info['attributes']['mail'])
    except KeyError:
        pass

    users[user_info['name']] = user_info
    user_info = {}

wb = load_workbook('users.xlsx')
ws = wb['main']
n = 1
i = str(n)

ws['A'+i] = 'Email'
ws['B'+i] = 'Name'
ws['C'+i] = 'Phone'
ws['D'+i] = 'Mobile'
ws['E'+i] = 'id'
ws['F'+i] = 'Job_Description'
ws['G'+i] = 'Manager_Name'
ws['H'+i] = 'Manager_Mail'
ws['I'+i] = 'Notes'
ws['J'+i] = 'Manager_id'

for user in users:
    n += 1
    i = str(n)
    for attribs in users[user]:
        ws['A'+i] = users[user]['mail']
        ws['B'+i] = users[user]['name']
        ws['C'+i] = ''
        try:
            ws['D'+i] = users[user]['mobile']
        except KeyError:
            ws['D'+i] = ''
        ws['E'+i] = users[user]['id']
        ws['F'+i] = users[user]['job_description']
        ws['G'+i] = users[user]['manager_name']
        ws['H'+i] = users[user]['manager_mail']
        ws['I'+i] = ''
        ws['J'+i] = users[user]['manager_id']

wb.save('%s.xlsx' % time.strftime("%d-%m-%Y_users"))
